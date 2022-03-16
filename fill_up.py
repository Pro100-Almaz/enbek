import logging
import xlrd
import openpyxl
from enbek import Enbek
import json
import config
import xlwings as xw
import Levenshtein
from datetime import datetime
import os
from xlutils.copy import copy as xl_copy

root_path = str(__file__).split("Sources")[0].replace("/", "\\")
root_path2 = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
files = root_path + "Files\\"
state = True

ignor_iin=[]
otpusk_done_iin = []
# f = open(r"C:\Users\Administrator\Desktop\RPA_Robot\RPA_Robot\Robot\EnbekRobot\Sources\iin.txt", "r")
# for x in f:
#   ignor_iin.append(str(x)[:12])
#
# f = open("C:\Users\Administrator\Desktop\RPA_Robot\RPA_Robot\Robot\EnbekRobot\Sources\iin_otpusk.txt", "r")
# for x in f:
#   otpusk_done_iin.append(str(x)[:12])



class Search_Data:
    mapping_files = []

    def __init__(self, name):
        self.name = name

    def define_filial(self):
        for variation in config.filials.keys():
            if variation == self.name:
                return config.filials[variation]

        return "none"


class Create_Data:
    def __init__(self, filial_name, account_enbek, password_enbek, path_ecp, key_ecp):

        # Отрываем страницу в браузере
        self.enbek = Enbek(root_path, account_enbek, password_enbek, path_ecp, key_ecp)

        mapping_files = root_path + "Files\\" + filial_name
        # Mapping
        map_adress = openpyxl.load_workbook(mapping_files + "\\Маппинг_адресов.xlsx")
        map_doljn = openpyxl.load_workbook(mapping_files + "\\Маппинг_должностей.xlsx")
        map_prichin = openpyxl.load_workbook(mapping_files + "\\Маппинг_причин.xlsx")
        self.mapping_address = map_adress["ЦФО_upd"]
        self.mapping_rastorg = map_prichin["Маппинг причин увольнений"]
        self.mapping_doljn = map_doljn["Маппинг_должностей"]

        # index 0: 503; index 1: 504; index 2: vacancies
        file1 = config.file_list[0] + filial_name + ".xls"
        # file1 = r"C:\Users\Administrator\Desktop\RPA_Robot\Robot\EnbekRobot\Files\test.xls"
        file2 = config.file_list[1] + filial_name + ".xlsm"
        file3 = config.file_list[2]

        # Для нумираций записей
        self.num = 1

        # Для договоров: 0 = Доб. труд. согл; 1 = Созд. доп. согл; 2 = Расторжение
        self.wb_dog = xlrd.open_workbook(file1)

        # Для отпуска: только один лист
        self.wb_otp = xlrd.open_workbook(file2)

        # Для вакансий: только один лист
        self.wb_vcns = openpyxl.load_workbook(file3)

        # Для сохранения результатов
        date = (str(datetime.today().date())).split("-")
        self.out_file_name = files + filial_name + "\\" + "RPA_Enbek_" + date[2] + date[1] + date[0] + ".xlsx"
        self.out_file_name_for_send = "RPA_Enbek_" + date[2] + date[1] + date[0] + ".xlsx"

    book = openpyxl.Workbook()
    book.create_sheet("Результаты")
    book.create_sheet("Соц. отпуска")
    book.create_sheet("Добавления труд. догов.")  # book["Добавления труд. догов."]
    book.create_sheet("Создание доп. согл.")  # book["Создание доп. согл."]
    book.create_sheet("Расторжения")  # book["Расторжения"]
    book.create_sheet("Вакансии")  # book["Вакансии"]
    ws = book["Результаты"]
    ws.append(['№', 'Дата', "ИИН", "Действие", "Статус", "Примечание"])

    def get_doljn(self, temp_address, name):
        sheet2 = self.mapping_doljn
        name = name.capitalize()

        restricting_sentences = ['Филиал акционерного общества «Национальная компания «Қазақстан теміржолы» - '
                                 '«Многофункциональный центр обслуживания»',
                                 'филиал акционерного общества «национальная '
                                 'компания «қазақстан теміржолы» - '
                                 '«многофункциональный центр обслуживания»',
                                 'представительства ао \"нк \"қтж\" за рубежом представительство ао \"нк \"қтж\"',
                                 'производственный персонал',
                                 'филиал акционерного общества "национальная компания \"қазақстан теміржолы\"',
                                 '\"павлодарское отделение магистральной сети\" филиал ао \"нк \"қтж\"',
                                 'филиал ао \"нк \"ктж\"',
                                 'производственный персонал',
                                 'Административный персонал',
                                 'Филиал акционерного общества \"Национальная компания \"Қазақстан теміржолы\" '
                                 '\"Центр оценки и развития персонала железнодорожного транспорта\"',
                                 'филиал акционерного общества «Национальная компания «Қазақстан теміржолы» – «Центр '
                                 'оценки и развития персонала железнодорожного транспорта»',
                                 'филиал АО \"НК \"КТЖ\" - \"Дирекция магистральной сети\" ',


                                 ]

        temp_address = temp_address.replace("учета", "учета станция")
        if temp_address == "":
            return "none"

        for rs_element in restricting_sentences:
            if temp_address is not None and Levenshtein.distance(rs_element, temp_address) / len(temp_address) < 0.5:
                temp_address = temp_address.replace(rs_element, '')
            temp_address = temp_address.strip()
            temp_address = temp_address.lower()
            temp_address = temp_address.replace("рцо", "региональный центр обслуживания")

            check = 0

            for row_num in range(1, sheet2.max_row + 1):
                temp = sheet2.cell(row=row_num, column=1).value
                if temp == None:
                    continue
                temp = temp.lower()
                temp = temp.strip()

                if temp != None and temp == temp_address:
                    check = 1

                if not check and len(temp_address) != 0:
                    accuracy = 1 - Levenshtein.distance(temp, temp_address) / len(temp_address)
                    if accuracy > 0.9:
                        check = 1
                if check:
                    if sheet2.cell(row=row_num, column=2).value == name:
                        if sheet2.cell(row=row_num, column=4).value == "бухгалтер":
                            return "Бухгалтер"
                        return sheet2.cell(row=row_num, column=4).value
        for row_num in range(1, sheet2.max_row + 1):
            if sheet2.cell(row=row_num, column=2).value == name:
                if sheet2.cell(row=row_num, column=4).value == "бухгалтер":
                    return "Бухгалтер"
                return sheet2.cell(row=row_num, column=4).value
        return "none"

    def get_address(self, full_address):
        out_list = []
        mapping_target = ""
        mapping_count = 0
        count = 0
        for i in range(len(full_address) - 1, 0, -1):
            if count == 2:
                break
            if full_address[i] == ',' and mapping_target != "":
                count += 1
            else:
                mapping_target += full_address[i]
                mapping_count += ord(full_address[i])
        for row in self.mapping_address.iter_rows():
            temp_address = ""
            check = 0
            temp_count = 0
            for i in row[2].value:
                if i == ',':
                    check = 1
                elif check:
                    temp_address += i
                    temp_count += ord(i)
            temp_val = mapping_target[::-1]
            temp_val.rstrip().lstrip()
            temp_address.rstrip().lstrip()
            if (temp_address[0:len(temp_address)] == temp_val[0:len(temp_address)] and len(temp_address) != 0) or mapping_count == temp_count:
                out_list.append(row[3].value)
                out_list.append(row[4].value)
                out_list.append(row[5].value)
                out_list.append(temp_address.lstrip())
                return out_list
        return "none"

    def get_adr(self, full_address):
        address = full_address.split(", ")
        return address[-1]

    def nomer_dog(self, num):
        if num == "Б/н":
            return num
        try:
            return int(num)
        except Exception as e:
            logging.debug("In function \"nomer_dog\", %s", e)

    def mapping_vidov_dog(self, str):
        if str == "На неопределенный срок":
            return str.lower()
        elif str == "На определенный срок":
            return "на определенный срок не менее одного года"
        else:
            return "на время замещения временно отсутствующего работника"

    def mapping_rastorg_func(self, str):
        sheet = self.mapping_rastorg
        for row_num in range(1, sheet.max_row + 1):
            if str == sheet.cell(row=row_num, column=1).value:
                return sheet.cell(row=row_num, column=2).value

        return "none"

    def create_sogl(self):
        sheet = self.wb_dog.sheet_by_index(0)
        date = datetime.now().strftime('%d.%m.%Y-%H:%M:%S')
        self.book["Добавления труд. догов."].append(["№", "Предприятие", "ИИН", "Номер договора", "Срок договора",
                                                     "Дата заключения договора", "Дата начала работы",
                                                     "Дата окончания действия договора  (зависит от \"Срок договора\")",
                                                     "Вид работы", "Подразделение Компании", "Штатная должность",
                                                     "Место выполнения работы (полный адрес)", "Режим рабочего времени",
                                                     "Военная обязанность"])
        for row_ind in range(1, sheet.nrows):
        # for row_ind in range(1, 9):

            if sheet.cell_value(row_ind, 2) != '' and str(int(sheet.cell_value(row_ind, 2))) in ignor_iin:
                continue

            self.book["Добавления труд. догов."].append([row_ind, str(sheet.cell_value(row_ind, 1)),
                                                         str(sheet.cell_value(row_ind, 2)),
                                                         str(self.nomer_dog(sheet.cell_value(row_ind, 3))),
                                                         str(sheet.cell_value(row_ind, 4)),
                                                         str(sheet.cell_value(row_ind, 5)),
                                                         str(sheet.cell_value(row_ind, 6)),
                                                         str(sheet.cell_value(row_ind, 7)),
                                                         str(sheet.cell_value(row_ind, 8)),
                                                         str(sheet.cell_value(row_ind, 9)),
                                                         str(sheet.cell_value(row_ind, 10)),
                                                         str(sheet.cell_value(row_ind, 11)),
                                                         str(sheet.cell_value(row_ind, 12)),
                                                         str(sheet.cell_value(row_ind, 13)),
                                                         ])

            address_arr = self.get_address(str(sheet.cell_value(row_ind, 11)))
            if sheet.cell_value(row_ind, 2) == '':
                self.ws.append([self.num, date, "Нету", "Создание трудового договора", "Не успешно",
                                "Отсутствие ИИН."])
                self.num += 1
                continue
            if sheet.cell_value(row_ind, 9) == '':
                self.ws.append([self.num, date, str(int(sheet.cell_value(row_ind, 2))), "Создание трудового договора", "Не успешно",
                                "Отсутствие Подразделение Компании."])
                self.num += 1
                continue
            if self.get_doljn(str(sheet.cell_value(row_ind, 9)), str(sheet.cell_value(row_ind, 10))) == "none":
                self.ws.append([self.num, date, str(int(sheet.cell_value(row_ind, 2))), "Создание трудового договора",
                                "Не успешно",
                                "Ошибка: не найден маппинг должности. Должность из SAP: " + str(
                                    sheet.cell_value(row_ind, 10))])
                self.num += 1
                continue

            if address_arr == "none":
                self.ws.append([self.num, date, str(int(sheet.cell_value(row_ind, 2))), "Создание трудового договора",
                                "Не успешно",
                                "Ошибка: не найден маппинг адреса. Адрес из SAP: " +
                                str(sheet.cell_value(row_ind, 11))])
                self.num += 1
                continue
            data_priem = {
                "ИИН": str(int(sheet.cell_value(row_ind, 2))),
                "Номер договора": str(self.nomer_dog(sheet.cell_value(row_ind, 3))),
                "Срок договора": self.mapping_vidov_dog(sheet.cell_value(row_ind, 4)),
                "Дата заключения договора": str(sheet.cell_value(row_ind, 5)),
                "Дата начала работы": str(sheet.cell_value(row_ind, 6)),
                "Дата окончания действия договора": str(sheet.cell_value(row_ind, 7)),
                "Штатная должность": str(sheet.cell_value(row_ind, 10)),
                "Вид работы": str(sheet.cell_value(row_ind, 8)).lower(),
                "Режим рабочего времени": str(sheet.cell_value(row_ind, 12)),
                "Военная обязанность": str(sheet.cell_value(row_ind, 13)).lower(),
                "Место выполнения работы": {
                    "obl": address_arr[0],
                    "city": address_arr[1],
                    "nas_punkt": address_arr[2],
                    "street": address_arr[3]
                },
                "Должность": self.get_doljn(str(sheet.cell_value(row_ind, 9)), str(sheet.cell_value(row_ind, 10)))
            }
            try:
                self.enbek.create_dogovor(data_priem)
                self.ws.append([self.num, date, data_priem["ИИН"], "Создание трудового договора", "Успешно"])
            except Exception as ve:
                logging.debug("FAIL--- %s", data_priem["ИИН"])
                if str(ve) == "No dogovor":
                    self.ws.append(
                        [self.num, date, data_priem["ИИН"], "Создание трудового договора", "Не успешно",
                         "Договор был создан ранее!"])
                elif str(ve) == "death":
                    self.ws.append(
                        [self.num, date, data_priem["ИИН"], "Создание трудового договора", "Не успешно",
                         "Человек с данным ИИН в государственной базе данных физических лиц значится умершим!"])
                elif str(ve) == "Clash in dogovor":
                    self.ws.append(
                        [self.num, date, data_priem["ИИН"], "Создание трудового договора", "Не успешно",
                         "Ошибка добавления! У вас уже есть действующий договор по данному ИИН."])
                elif str(ve) == "Tricky dogovor!":
                    self.ws.append(
                        [self.num, date, data_priem["ИИН"], "Создание трудового договора", "Не успешно",
                         "Ошибка добавления! Имеется уже действующяя операция!"])
                else:
                    self.ws.append(
                        [self.num, date, data_priem["ИИН"], "Создание трудового договора", "Не успешно",
                         "Ошибка: сбой системы. Текст ошибки: " + str(ve)])
            self.num += 1
            self.book.save(self.out_file_name)


    def dop_sogl(self):
        sheet = self.wb_dog.sheet_by_index(1)
        date = datetime.now().strftime('%d.%m.%Y-%H:%M:%S')
        self.book["Создание доп. согл."].append(["№", "Предприятие", "ИИН", "Номер дополнительного соглашения",
                                                 "Срок договора", "Вид работы",
                                                 "Дата окончания действия договора  (зависит от \"Срок договора\")",
                                                 "Дата заключения дополнительного соглашения",
                                                 "Дата начала действия дополнительного соглашения",
                                                 "Дата окончания действия дополнительного соглашения",
                                                 "Дата начала работы",
                                                 "Подразделение Компании", "Штатная должность",
                                                 "Место выполнения работы (полный адрес)", "Режим рабочего времени",
                                                 ])
        for row_ind in range(1, sheet.nrows):
        # for row_ind in range(1, 3):

            self.book["Создание доп. согл."].append([row_ind, str(sheet.cell_value(row_ind, 1)),
                                                     str(sheet.cell_value(row_ind, 2)),
                                                     str(sheet.cell_value(row_ind, 3)),
                                                     str(sheet.cell_value(row_ind, 4)),
                                                     str(sheet.cell_value(row_ind, 5)),
                                                     str(sheet.cell_value(row_ind, 6)),
                                                     str(sheet.cell_value(row_ind, 7)),
                                                     str(sheet.cell_value(row_ind, 8)),
                                                     str(sheet.cell_value(row_ind, 9)),
                                                     str(sheet.cell_value(row_ind, 10)),
                                                     str(sheet.cell_value(row_ind, 11)),
                                                     str(sheet.cell_value(row_ind, 12)),
                                                     str(sheet.cell_value(row_ind, 13)),
                                                     str(sheet.cell_value(row_ind, 14)),
                                                     ])

            address_arr = self.get_address(str(sheet.cell_value(row_ind, 13)))
            if self.mapping_vidov_dog(sheet.cell_value(row_ind, 4)) == "на неопределенный срок":
                data_end_of_dog = ""
            else:
                data_end_of_dog = str(sheet.cell_value(row_ind, 6))

            if sheet.cell_value(row_ind, 2) == '':
                self.ws.append([self.num, date, "Нету", "Дополнительное соглашение", "Не успешно",
                                "Отсутствие ИИН."])
                self.num += 1
                continue
            if self.get_doljn(str(sheet.cell_value(row_ind, 11)), str(sheet.cell_value(row_ind, 12))) == "none":
                self.ws.append([self.num, date, str(int(sheet.cell_value(row_ind, 2))), "Дополнительное соглашение",
                                "Не успешно",
                                "Ошибка: не найден маппинг должности. Должность из SAP: " + str(
                                    sheet.cell_value(row_ind, 12))])
                self.num += 1
                continue
            if address_arr == "none":
                self.ws.append([self.num, date, str(int(sheet.cell_value(row_ind, 2))), "Дополнительное соглашение",
                                "Не успешно",
                                "Ошибка: не найден маппинг адреса. Адрес из SAP: " +
                                str(sheet.cell_value(row_ind, 11))])
                self.num += 1
                continue

            data_dop_sogl = {
                "ИИН": str(int(sheet.cell_value(row_ind, 2))),
                "Номер доп соглашения": str(self.nomer_dog(sheet.cell_value(row_ind, 3))),
                "Срок договора": self.mapping_vidov_dog(sheet.cell_value(row_ind, 4)),
                "Дата окончания действия трудового договора": data_end_of_dog,
                "Дата начала действия доп соглашения": str(sheet.cell_value(row_ind, 8)),
                "Дата заключения дополнительного соглашения": str(sheet.cell_value(row_ind, 7)),
                "Дата окончания действия дополнительного соглашения": str(sheet.cell_value(row_ind, 9)),
                "Дата начала работы": str(sheet.cell_value(row_ind, 10)),
                "Штатная должность": str(sheet.cell_value(row_ind, 12)),
                "Вид работы": str(sheet.cell_value(row_ind, 5)).lower(),
                "Режим рабочего времени": str(sheet.cell_value(row_ind, 14)),
                "Военная обязанность": "призывник",
                "Место выполнения работы": {
                    "obl": address_arr[0],
                    "city": address_arr[1],
                    "nas_punkt": address_arr[2],
                    "street": address_arr[3]
                },
                "Должность": self.get_doljn(str(sheet.cell_value(row_ind, 11)), str(sheet.cell_value(row_ind, 12)))
            }

            try:
                self.enbek.create_dop_sogl(data_dop_sogl)
                self.ws.append([self.num, date, data_dop_sogl["ИИН"], "Дополнительное соглашение", "Успешно"])
            except Exception as e:
                logging.debug("FAIL--- %s", data_dop_sogl["ИИН"])
                if str(e) == "Доп. соглашение уже создано":
                    self.ws.append([self.num, date, data_dop_sogl["ИИН"], "Дополнительное соглашение", "Не успешно",
                                    "Такое дополнительное соглашение уже существует!"])
                elif str(e) == "No dogovor":
                    self.ws.append([self.num, date, data_dop_sogl["ИИН"], "Дополнительное соглашение", "Не успешно",
                                    "Отсутствие договора!"])
                else:
                    self.ws.append([self.num, date, data_dop_sogl["ИИН"], "Дополнительное соглашение", "Не успешно",
                                    "Ошибка: сбой системы. Текст ошибки: " + str(e)])
            self.num += 1
            self.book.save(self.out_file_name)

    def delete_dog(self):
        date = datetime.now().strftime('%d.%m.%Y-%H:%M:%S')
        sheet = self.wb_dog.sheet_by_index(2)
        self.book["Расторжения"].append(["№", "Предприятие", "ИИН", "Дата расторжения договора",
                                         "Причина расторжения"])
        for row_ind in range(1, sheet.nrows):
        # for row_ind in range(10, 11):

            if sheet.cell_value(row_ind, 2) != '' and str(int(sheet.cell_value(row_ind, 2))) in ignor_iin:
                continue

            self.book["Расторжения"].append([row_ind, str(sheet.cell_value(row_ind, 1)),
                                             str(sheet.cell_value(row_ind, 2)),
                                             str(sheet.cell_value(row_ind, 3)),
                                             str(sheet.cell_value(row_ind, 4))])

            if sheet.cell_value(row_ind, 2) == '':
                self.ws.append([self.num, date, "Нету", "Расторжение", "Не успешно",
                                "Отсутствие ИИН."])
                self.num += 1
                continue

            if sheet.cell_value(row_ind, 3) == '':
                self.ws.append([self.num, date, str(int(sheet.cell_value(row_ind, 2))), "Расторжение", "Не успешно",
                                "Отсутствие даты расторжения."])
                self.num += 1
                continue

            if self.mapping_rastorg_func(str(sheet.cell_value(rowx=row_ind, colx=4))) == "none":
                self.ws.append([self.num, date, str(int(sheet.cell_value(row_ind, 2))), "Расторжение",
                                "Не успешно",
                                "Ошибка: не найден маппинг причины расторжения. Причина из SAP: " +
                                str(sheet.cell_value(row_ind, 4))])
                self.num += 1
                continue

            data_udalenie = {
                "ИИН": str(int(sheet.cell_value(row_ind, 2))),
                "Причина": self.mapping_rastorg_func(str(sheet.cell_value(row_ind, 4))),
                "Дата расторжение": str(sheet.cell_value(row_ind, 3)),
            }

            try:
                self.enbek.terminate_dog(data_udalenie)
                self.ws.append([self.num, date, data_udalenie["ИИН"], "Расторжение", "Успешно"])
            except Exception as e:
                logging.debug("FAIL--- %s", data_udalenie["ИИН"])
                if str(e) == "No dogovor":
                    self.ws.append([self.num, date, data_udalenie["ИИН"], "Расторжение", "Не успешно",
                                    "Сотрудник не найден!"])
                else:
                    self.ws.append([self.num, date, data_udalenie["ИИН"], "Расторжение", "Не успешно",
                                    "Ошибка: сбой системы. Текст ошибки: " + str(e)])
            self.num += 1
            self.book.save(self.out_file_name)

    def otpusk(self):
        self.book["Соц. отпуска"].append(["№", "Предприятие", "ИИН", "Тип отпуска", "Не работал \"С\"",
                                          "Не работал \"По\"", "Номер табеля (заполняется при случае если выбран тип "
                                                               "отпуска в связи с беременностью)",
                                          "Выходные дни за период нетрудоспособности (заполняется при случае если "
                                          "выбран тип отпуска в связи с беременностью)",
                                          "Наименование отдела на русском языке (заполняется "
                                          "при случае если выбран тип отпуска в связи с беременностью) ",
                                          "Наименование отдела на казахском язык (заполняется при случае если выбран "
                                          "тип отпуска в связи с беременностью)"])
        for sheet in self.wb_otp.sheets():
            if sheet.name == "Социальный отпуск":
                date = datetime.now().strftime('%d.%m.%Y-%H:%M:%S')

                for row in range(1, sheet.nrows):
                # for row in range(1, 273):

                    cell = sheet.row_values(row)
                    if type(cell[2]) != float:
                        continue

                    if str(int(cell[2])) in otpusk_done_iin:
                        continue

                    self.book["Соц. отпуска"].append([row, cell[1], int(cell[2]), cell[3], cell[4], cell[5], cell[6]])
                    num_tabel = ''
                    if cell[6]:
                        num_tabel = int(cell[6])

                    if cell[2] == '':
                        self.ws.append([self.num, date, "Нету", "Создание трудового договора", "Не успешно",
                                        "Отсутствие ИИН."])
                        self.num += 1
                        continue

                    if str(cell[
                               3]) == "В связи с беременностью и рождением ребенка, усыновления/удочерения новорожденного ребенка (детей)":
                        right_form = " В связи с беременностью и рождением ребенка, усыновления/удочерения новорожденного ребенка (детей)"
                    else:
                        right_form = str(cell[3])

                    data_otpusk = {
                        "ИИН": int(cell[2]),
                        "Тип отпуска": right_form,
                        "Дата с": str(cell[4]),
                        "Дата по": str(cell[5]),
                        "Номер табеля": num_tabel,
                        "дата первого рабочего дня": str(cell[5]),
                    }

                    try:
                        self.enbek.to_otpusk(data_otpusk)
                        self.ws.append([self.num, date, data_otpusk["ИИН"], "Социальный отпуск", "Успешно"])
                    except Exception as e:
                        logging.debug("FAIL--- %s", data_otpusk["ИИН"])
                        if str(e) == "No dogovor":
                            self.ws.append([self.num, date, data_otpusk["ИИН"], "Социальный отпуск", "Не успешно",
                                    "Сотрудник не найден!"])
                        elif str(e) == "Соц. отпуск уже создано":
                            self.ws.append([self.num, date, data_otpusk["ИИН"], "Социальный отпуск", "Не успешно",
                                    "Такой социальный отпуск уже существует!"])
                        elif str(e) == "Soc s datoi uje est":
                            self.ws.append([self.num, date, data_otpusk["ИИН"], "Социальный отпуск", "Не успешно",
                                            "Социальный отпуск с указаной датой к данному трудовому договору уже существует."])
                        elif str(e) == "Soc s datoi uje est and delete sohranenie":
                            self.ws.append([self.num, date, data_otpusk["ИИН"], "Социальный отпуск", "Не успешно",
                                        "Социальный отпуск с указаной датой к данному трудовому договору уже "
                                        "существует, также надо удалить сохранение."])
                        elif str(e) == "Soc s datoi posle sohraneniya":
                            self.ws.append([self.num, date, data_otpusk["ИИН"], "Социальный отпуск", "Не успешно",
                                            "Дата \"Не работал с\" не должна быть раньше даты начала договора, "
                                            "также надо удалить сохранение."
                                            ])
                        elif str(e) == "Already podpisan!":
                            self.ws.append([self.num, date, data_otpusk["ИИН"], "Социальный отпуск", "Не успешно",
                                            "Договор был подписан!"
                                            ])
                        else:
                            self.ws.append([self.num, date, data_otpusk["ИИН"], "Социальный отпуск", "Не успешно",
                                            "Ошибка: сбой системы. Текст ошибки: " + str(e)])
                    self.num += 1
                    self.book.save(self.out_file_name)

    def vacancies(self):
        date = datetime.now().strftime('%d.%m.%Y-%H:%M:%S')
        data = self.wb_vcns["Вакансии"]
        self.book["Соц. отпуска"].append(
            ["№", "", "Штатная должность (должно соответствовать колонке во вкладке \"Должности\")",
             "Уточнение должности", "Область", "Район или город", "Населенный пункт",
             "Полный адрес", "Режим работы", "Характер работы", "Условия труда",
             "Стажировка", "Оплата труда(тенге)- ОТ (не менее 42500 тг)",
             "Оплата труда(тенге)- ДО (опционально)", "Вакансия для выпускника ВУЗа",
             "Общее кол вакантных мест", "Из них молодежи (указать количество)",
             "Из них для лиц с инвалидностью (указать количество)",
             "Из них для состоящих на учете службы пробации (указать количество)",
             "Из них для освобожденных из мест лишения свободы (указать количество)",
             "Из них для молодежи (до совершеннолетия) без попечения родителей (указать количество)",
             "Дополнительные условия, социальная поддержка, в выборе где (тенге, гектар, кв.м.) нужно указать количество",
             "Стаж работы по специальности", "Профессиональные навыки (можно перечислить несколько через ,)",
             "Личные качества (можно перечислить несколько через ,)", "Уровень образования",
             "Должностные обязанности (текст)", "Водительские права (категории) (можно перечислить несколько через ,)",
             "Владение языками (Язык)", "Владение языками (Уровень владения)",
             "Срок публикации вакансии на портале"
             ])
        for i in range(2, data.max_row + 1):
            if data.cell(row=i, column=13).value is None:
                break
            # self.book["Создание доп. согл."].append([row_ind, data.cell(row=i, column=2).value,
            #                                          data.cell(row=i, column=3).value,
            #                                          data.cell(row=i, column=4).value,
            #                                          data.cell(row=i, column=5).value,
            #                                          data.cell(row=i, column=6).value,
            #                                          data.cell(row=i, column=7).value,
            #                                          data.cell(row=i, column=8).value,
            #                                          data.cell(row=i, column=9).value,
            #                                          data.cell(row=i, column=10).value,
            #                                          data.cell(row=i, column=11).value,
            #                                          data.cell(row=i, column=12).value,
            #                                          data.cell(row=i, column=13).value,
            #                                          data.cell(row=i, column=14).value,
            #                                          data.cell(row=i, column=15).value,
            #                                          data.cell(row=i, column=16).value,
            #                                          data.cell(row=i, column=17).value,
            #                                          data.cell(row=i, column=18).value,
            #                                          data.cell(row=i, column=19).value,
            #                                          data.cell(row=i, column=20).value,
            #                                          data.cell(row=i, column=21).value,
            #                                          data.cell(row=i, column=22).value,
            #                                          data.cell(row=i, column=23).value,
            #                                          data.cell(row=i, column=24).value,
            #                                          data.cell(row=i, column=25).value,
            #                                          data.cell(row=i, column=26).value,
            #                                          data.cell(row=i, column=27).value,
            #                                          data.cell(row=i, column=28).value,
            #                                          data.cell(row=i, column=29).value,
            #                                          data.cell(row=i, column=30).value,
            #                                          data.cell(row=i, column=31).value,
            #                                          ])
            data_vcns = {
                'Профессиональная область': data.cell(row=i, column=2).value,
                'Наименование должности': data.cell(row=i, column=3).value,
                'Уточнение должности': data.cell(row=i, column=4).value,
                'Адрес': {
                    'obl': (data.cell(row=i, column=5).value).upper(),
                    'center': data.cell(row=i, column=6).value,
                    'nas_punkt': data.cell(row=i, column=7).value,
                    'adres': data.cell(row=i, column=8).value,
                },
                'Режим работы': data.cell(row=i, column=9).value,
                'Характер работы': data.cell(row=i, column=10).value,
                'Условия труда': data.cell(row=i, column=11).value,
                'Стажировка': data.cell(row=i, column=12).value,
                'Оплата труда (тенге)': int(str(data.cell(row=i, column=13).value)),  # might be a error
                'Вакантные места': data.cell(row=i, column=16).value,
                'Стаж работы по специальности': data.cell(row=i, column=23).value,
                'Профессиональные навыки': str(data.cell(row=i, column=24).value).split(", "),
                'Личные качества': (data.cell(row=i, column=25).value).split(),
                'Уровень образования': data.cell(row=i, column=26).value,
                'Должностные обязанности': data.cell(row=i, column=27).value,
                'Срок публикации вакансии на портале': data.cell(row=i, column=31).value
            }

            try:
                self.enbek.add_vac(data_vcns)
                self.ws.append([self.num, date, data_vcns["Наименование должности"], "Вакансии", "Успешно"])
            except Exception as e:
                print(e)
                logging.debug("FAIL--- %s", data_vcns["Наименование должности"])
                self.ws.append(
                    [self.num, date, data_vcns["Наименование должности"], "Вакансии", "Не успешно",
                     "Неправильно заполнены данные!"])
            self.num += 1
            self.book.save(self.out_file_name)


    def saved_file_data(self):
        return [self.out_file_name, self.out_file_name_for_send]